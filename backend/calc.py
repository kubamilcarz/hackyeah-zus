"""
Kalkulator Emerytalny ZUS - Backend Implementation
Kompletna logika kalkulacji emerytury zgodna z polskim systemem emerytalnym
"""

import math
import os
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from enum import Enum
import numpy as np
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


class Gender(Enum):
    MALE = "M"
    FEMALE = "K"


@dataclass
class PensionParameters:
    """Parametry makroekonomiczne i systemowe"""
    # Waloryzacja składek
    indexation_rates: Dict[int, float]  # rok -> wskaźnik waloryzacji
    # Prognozowane średnie wynagrodzenia
    average_salaries: Dict[int, float]  # rok -> średnia krajowa
    # Inflacja
    inflation_rates: Dict[int, float]  # rok -> stopa inflacji
    # Wiek emerytalny
    retirement_age_male: float = 65.0
    retirement_age_female: float = 60.0
    # Minimalne świadczenie
    minimum_pension: float = 1588.44  # stan na 2025
    # Średnie dalsze trwanie życia (GUS)
    life_expectancy_at_retirement: Dict[tuple, float] = None  # (płeć, wiek) -> miesiące
    # Wskaźnik absencji chorobowej
    sick_leave_avg_days_male: float = 15.2  # dni/rok
    sick_leave_avg_days_female: float = 21.5  # dni/rok

    def __post_init__(self):
        if self.life_expectancy_at_retirement is None:
            # Tablice trwania życia GUS (przykładowe wartości)
            self.life_expectancy_at_retirement = {
                (Gender.MALE, 60): 240.5,  # 20 lat
                (Gender.MALE, 65): 192.8,  # 16 lat
                (Gender.MALE, 67): 176.4,  # 14.7 lat
                (Gender.FEMALE, 60): 276.3,  # 23 lata
                (Gender.FEMALE, 65): 234.6,  # 19.5 lat
                (Gender.FEMALE, 67): 216.8,  # 18 lat
            }


@dataclass
class UserInput:
    """Dane wprowadzane przez użytkownika"""
    # Dane podstawowe
    age: int
    gender: Gender
    gross_salary: float  # obecne wynagrodzenie brutto

    # Historia pracy
    work_start_year: int
    planned_retirement_year: int

    # Opcjonalne
    account_balance: Optional[float] = None  # stan konta w ZUS
    subaccount_balance: Optional[float] = None  # stan subkonta w ZUS
    include_sick_leave: bool = True  # uwzględniać absencję chorobową

    # Dodatkowe okresy składkowe
    additional_contribution_years: float = 0  # np. służba wojskowa, urlop wychowawczy

    # Historia wynagrodzeń (opcjonalne - dla dokładniejszych obliczeń)
    salary_history: Optional[Dict[int, float]] = None  # rok -> wynagrodzenie

    # Kod pocztowy (do statystyk)
    postal_code: Optional[str] = None

    # PPK/IKZE (III filar)
    ppk_contribution_rate: float = 0.02  # 2% podstawowa składka pracownika
    ppk_employer_rate: float = 0.015  # 1.5% składka pracodawcy
    ikze_annual_amount: float = 0  # roczna wpłata na IKZE


class PensionCalculator:
    """Główny kalkulator emerytalny"""

    def __init__(self, parameters: PensionParameters):
        self.params = parameters
        self.current_year = datetime.now().year

    def calculate(self, user_input: UserInput) -> Dict:
        """Główna metoda kalkulacji emerytury"""

        # 1. Oblicz kapitał początkowy (dla osób pracujących przed 1999)
        initial_capital = self._calculate_initial_capital(user_input)

        # 2. Oblicz zgromadzony kapitał z okresu po 1999
        accumulated_capital = self._calculate_accumulated_capital(user_input)

        # 3. Oblicz prognozowany kapitał do emerytury
        future_capital = self._calculate_future_capital(user_input)

        # 4. Oblicz średnie dalsze trwanie życia
        life_expectancy = self._get_life_expectancy(
            user_input.gender,
            user_input.planned_retirement_year - user_input.work_start_year +
            self.current_year - user_input.age
        )

        # 5. Oblicz emeryturę
        total_capital = initial_capital + accumulated_capital + future_capital

        # Podstawowa emerytura (kapitał / średnie dalsze trwanie życia w miesiącach)
        base_pension = total_capital / life_expectancy

        # 6. Sprawdź minimalną emeryturę
        years_worked = user_input.planned_retirement_year - user_input.work_start_year
        min_years_required = 25 if user_input.gender == Gender.MALE else 20

        if years_worked >= min_years_required:
            base_pension = max(base_pension, self.params.minimum_pension)

        # 7. Oblicz emeryturę urealnioną (z uwzględnieniem inflacji)
        real_pension = self._calculate_real_value(
            base_pension,
            user_input.planned_retirement_year
        )

        # 8. Oblicz stopę zastąpienia
        final_salary = self._estimate_final_salary(user_input)
        replacement_rate = (base_pension / final_salary) * 100 if final_salary > 0 else 0

        # 9. Symulacje różnych scenariuszy
        scenarios = self._calculate_scenarios(user_input, total_capital)

        # 10. Porównanie ze średnią krajową
        avg_pension = self._get_average_pension(user_input.planned_retirement_year)
        comparison_to_avg = (base_pension / avg_pension) * 100 if avg_pension > 0 else 100

        # 11. PPK/IKZE
        third_pillar = self._calculate_third_pillar(user_input)

        return {
            "pension_nominal": round(base_pension, 2),
            "pension_real": round(real_pension, 2),
            "total_capital": round(total_capital, 2),
            "initial_capital": round(initial_capital, 2),
            "accumulated_capital": round(accumulated_capital, 2),
            "future_capital": round(future_capital, 2),
            "replacement_rate": round(replacement_rate, 2),
            "comparison_to_average": round(comparison_to_avg, 2),
            "average_pension": round(avg_pension, 2),
            "life_expectancy_months": round(life_expectancy, 1),
            "years_worked": years_worked,
            "retirement_year": user_input.planned_retirement_year,
            "sick_leave_impact": self._calculate_sick_leave_impact(user_input),
            "scenarios": scenarios,
            "third_pillar": third_pillar,
            "meets_minimum_requirements": years_worked >= min_years_required
        }

    def _calculate_initial_capital(self, user_input: UserInput) -> float:
        """Oblicz kapitał początkowy dla lat pracy przed 1999"""
        if user_input.work_start_year >= 1999:
            return 0.0

        years_before_1999 = min(1999 - user_input.work_start_year,
                                1999 - (self.current_year - user_input.age - 18))

        if years_before_1999 <= 0:
            return 0.0

        # Uproszczony wzór - w rzeczywistości bardziej skomplikowany
        # Bazuje na średniej podstawie wymiaru składek z 10 lat
        base_amount = user_input.gross_salary * 0.7  # szacunkowa podstawa historyczna

        # Wzór: 24% podstawy wymiaru × liczba lat × 12 miesięcy × współczynnik
        capital = base_amount * 0.24 * years_before_1999 * 12 * 1.3

        # Waloryzacja kapitału początkowego
        for year in range(1999, self.current_year + 1):
            if year in self.params.indexation_rates:
                capital *= (1 + self.params.indexation_rates[year])

        return capital

    def _calculate_accumulated_capital(self, user_input: UserInput) -> float:
        """Oblicz dotychczas zgromadzony kapitał (konto + subkonto)"""

        # Jeśli podano wartości kont, użyj ich
        if user_input.account_balance is not None:
            return user_input.account_balance + (user_input.subaccount_balance or 0)

        # W przeciwnym razie oszacuj na podstawie historii wynagrodzeń
        accumulated = 0.0
        start_year = max(user_input.work_start_year, 1999)

        for year in range(start_year, self.current_year + 1):
            # Pobierz wynagrodzenie z historii lub oszacuj
            if user_input.salary_history and year in user_input.salary_history:
                yearly_salary = user_input.salary_history[year]
            else:
                # Oszacuj historyczne wynagrodzenie
                yearly_salary = self._estimate_historical_salary(
                    user_input.gross_salary,
                    self.current_year - year
                )

            # Składka emerytalna: 19.52% (9.76% pracownik + 9.76% pracodawca)
            yearly_contribution = yearly_salary * 12 * 0.1952

            # Podział na konto i subkonto (od 2014 roku)
            if year < 2014:
                # Całość na konto
                account_contribution = yearly_contribution
                subaccount_contribution = 0
            else:
                # 12.22% na konto, 7.3% na subkonto (przed 2017)
                # Później 15.52% na konto, 4% na subkonto
                if year < 2017:
                    account_contribution = yearly_salary * 12 * 0.1222
                    subaccount_contribution = yearly_salary * 12 * 0.073
                else:
                    account_contribution = yearly_salary * 12 * 0.1552
                    subaccount_contribution = yearly_salary * 12 * 0.04

            # Waloryzacja konta głównego
            for val_year in range(year + 1, self.current_year + 1):
                if val_year in self.params.indexation_rates:
                    account_contribution *= (1 + self.params.indexation_rates[val_year])

            # Subkonto - wzrost jak w OFE (średnio 5% rocznie)
            subaccount_growth = pow(1.05, self.current_year - year)
            subaccount_contribution *= subaccount_growth

            accumulated += account_contribution + subaccount_contribution

        # Uwzględnij absencję chorobową
        if user_input.include_sick_leave:
            sick_days = (self.params.sick_leave_avg_days_female
                         if user_input.gender == Gender.FEMALE
                         else self.params.sick_leave_avg_days_male)
            reduction = sick_days / 365 * 0.2  # 80% podstawy podczas choroby
            accumulated *= (1 - reduction)

        return accumulated

    def _calculate_future_capital(self, user_input: UserInput) -> float:
        """Prognozuj kapitał do emerytury"""
        future = 0.0

        for year in range(self.current_year + 1, user_input.planned_retirement_year + 1):
            # Prognozowane wynagrodzenie
            years_ahead = year - self.current_year
            salary_growth = pow(1.04, years_ahead)  # średni wzrost 4% rocznie
            projected_salary = user_input.gross_salary * salary_growth

            # Składka roczna
            yearly_contribution = projected_salary * 12 * 0.1952

            # Podział konto/subkonto
            account_contribution = projected_salary * 12 * 0.1552
            subaccount_contribution = projected_salary * 12 * 0.04

            # Waloryzacja do emerytury
            years_to_retirement = user_input.planned_retirement_year - year

            # Konto główne - waloryzacja średnio 3% rocznie
            account_future = account_contribution * pow(1.03, years_to_retirement)

            # Subkonto - wzrost 5% rocznie
            subaccount_future = subaccount_contribution * pow(1.05, years_to_retirement)

            future += account_future + subaccount_future

        # Uwzględnij absencję
        if user_input.include_sick_leave:
            sick_days = (self.params.sick_leave_avg_days_female
                         if user_input.gender == Gender.FEMALE
                         else self.params.sick_leave_avg_days_male)
            reduction = sick_days / 365 * 0.2
            future *= (1 - reduction)

        return future

    def _get_life_expectancy(self, gender: Gender, retirement_age: int) -> float:
        """Pobierz średnie dalsze trwanie życia"""
        # Znajdź najbliższy wiek z tablic
        age_key = min(self.params.life_expectancy_at_retirement.keys(),
                      key=lambda x: abs(x[1] - retirement_age) if x[0] == gender else float('inf'))

        if age_key[0] != gender:
            # Użyj wartości domyślnej
            return 240.0 if gender == Gender.FEMALE else 200.0

        return self.params.life_expectancy_at_retirement[age_key]

    def _calculate_real_value(self, nominal_value: float, target_year: int) -> float:
        """Przelicz wartość nominalną na realną (z uwzględnieniem inflacji)"""
        if target_year <= self.current_year:
            return nominal_value

        cumulative = 1.0
        last_rate = 0.025
        for year in range(self.current_year + 1, target_year + 1):
            rate = self.params.inflation_rates.get(year, last_rate) if hasattr(self.params, 'inflation_rates') and self.params.inflation_rates else last_rate
            if isinstance(rate, (int, float)):
                last_rate = float(rate)
            cumulative *= (1.0 + float(last_rate))
        real_value = nominal_value / cumulative
        return real_value

    def _estimate_final_salary(self, user_input: UserInput) -> float:
        """Oszacuj ostatnie wynagrodzenie przed emeryturą"""
        years_to_retirement = user_input.planned_retirement_year - self.current_year

        salary_growth = None
        if hasattr(self.params, 'average_salaries') and self.params.average_salaries:
            current_avg = self.params.average_salaries.get(self.current_year)
            target_avg = self.params.average_salaries.get(user_input.planned_retirement_year)
            if current_avg and target_avg and current_avg > 0:
                salary_growth = float(target_avg) / float(current_avg)
        if salary_growth is None:
            # Zakładamy średni wzrost wynagrodzeń 4% rocznie
            salary_growth = pow(1.04, max(0, years_to_retirement))
        return user_input.gross_salary * salary_growth

    def _estimate_historical_salary(self, current_salary: float, years_ago: int) -> float:
        """Oszacuj wynagrodzenie sprzed X lat"""
        if years_ago <= 0:
            return current_salary
        if hasattr(self.params, 'average_salaries') and self.params.average_salaries:
            past_year = self.current_year - years_ago
            current_avg = self.params.average_salaries.get(self.current_year)
            past_avg = self.params.average_salaries.get(past_year)
            if current_avg and past_avg and current_avg > 0:
                factor = float(past_avg) / float(current_avg)
                return current_salary * factor
        # Fallback: Zakładamy średni wzrost 4% rocznie
        return current_salary / pow(1.04, years_ago)

    def _get_average_pension(self, year: int) -> float:
        """Pobierz prognozowaną średnią emeryturę"""
        # Obecna średnia to około 3200 zł, wzrost 5% rocznie
        current_average = 3200.0
        years_ahead = year - self.current_year
        return current_average * pow(1.05, years_ahead)

    def _calculate_sick_leave_impact(self, user_input: UserInput) -> Dict:
        """Oblicz wpływ absencji chorobowej"""
        if not user_input.include_sick_leave:
            return {"reduction_percent": 0, "reduction_amount": 0}

        sick_days = (self.params.sick_leave_avg_days_female
                     if user_input.gender == Gender.FEMALE
                     else self.params.sick_leave_avg_days_male)

        # Podczas choroby podstawa składek to 80% wynagrodzenia
        years_worked = user_input.planned_retirement_year - user_input.work_start_year
        total_sick_days = sick_days * years_worked

        # Wpływ na emeryturę (uproszczony)
        reduction_percent = (total_sick_days / (years_worked * 365)) * 20  # 20% różnicy w składkach

        # Oblicz bez chorobowego
        calc_without_sick = self.calculate({**user_input.__dict__, 'include_sick_leave': False})
        reduction_amount = calc_without_sick["pension_nominal"] - \
                           self.calculate(user_input)["pension_nominal"]

        return {
            "total_sick_days": round(total_sick_days),
            "reduction_percent": round(reduction_percent, 2),
            "reduction_amount": round(reduction_amount, 2),
            "average_days_per_year": round(sick_days, 1)
        }

    def _calculate_scenarios(self, user_input: UserInput, base_capital: float) -> Dict:
        """Oblicz różne scenariusze emerytalne"""
        scenarios = {}

        # Scenariusz: opóźnienie emerytury o 1, 2, 5 lat
        for delay in [1, 2, 5]:
            delayed_input = UserInput(
                age=user_input.age,
                gender=user_input.gender,
                gross_salary=user_input.gross_salary,
                work_start_year=user_input.work_start_year,
                planned_retirement_year=user_input.planned_retirement_year + delay,
                account_balance=user_input.account_balance,
                subaccount_balance=user_input.subaccount_balance,
                include_sick_leave=user_input.include_sick_leave
            )

            delayed_result = self.calculate(delayed_input)

            scenarios[f"delay_{delay}_year"] = {
                "pension": delayed_result["pension_nominal"],
                "increase_amount": delayed_result["pension_nominal"] - \
                                   self.calculate(user_input)["pension_nominal"],
                "increase_percent": ((delayed_result["pension_nominal"] /
                                      self.calculate(user_input)["pension_nominal"]) - 1) * 100
            }

        # Scenariusz: ile lat pracy dla oczekiwanej emerytury
        # (to wymaga iteracji - uproszczona wersja)

        return scenarios

    def _calculate_third_pillar(self, user_input: UserInput) -> Dict:
        """Oblicz dodatkowe oszczędności emerytalne (PPK, IKZE)"""

        years_to_retirement = user_input.planned_retirement_year - self.current_year

        # PPK
        ppk_total = 0
        if user_input.ppk_contribution_rate > 0:
            yearly_ppk = user_input.gross_salary * 12 * (
                    user_input.ppk_contribution_rate + user_input.ppk_employer_rate
            )

            # Zakładamy wzrost 6% rocznie (fundusze)
            for year in range(years_to_retirement):
                ppk_total += yearly_ppk * pow(1.06, years_to_retirement - year)

        # IKZE
        ikze_total = 0
        if user_input.ikze_annual_amount > 0:
            for year in range(years_to_retirement):
                ikze_total += user_input.ikze_annual_amount * pow(1.06, years_to_retirement - year)

        # Przelicz na miesięczną rentę
        life_expectancy = self._get_life_expectancy(
            user_input.gender,
            user_input.planned_retirement_year - user_input.work_start_year +
            self.current_year - user_input.age
        )

        ppk_monthly = ppk_total / life_expectancy if ppk_total > 0 else 0
        ikze_monthly = ikze_total / life_expectancy if ikze_total > 0 else 0

        return {
            "ppk_total": round(ppk_total, 2),
            "ppk_monthly": round(ppk_monthly, 2),
            "ikze_total": round(ikze_total, 2),
            "ikze_monthly": round(ikze_monthly, 2),
            "total_additional": round(ppk_total + ikze_total, 2),
            "monthly_additional": round(ppk_monthly + ikze_monthly, 2)
        }


# Funkcja pomocnicza do inicjalizacji kalkulatora z domyślnymi parametrami
# === Loader parametrów z pliku Excel ===

def _parse_year_value_sheet(ws) -> Dict[int, float]:
    data: Dict[int, float] = {}
    try:
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            year = row[0]
            value = None
            # find first numeric after year
            for cell in row[1:]:
                if isinstance(cell, (int, float)):
                    value = float(cell)
                    break
            if isinstance(year, (int, float)):
                y = int(year)
                if 1900 < y < 2101 and value is not None:
                    data[y] = float(value)
    except Exception:
        return {}
    return data


def load_parameters_from_excel(path: str) -> Optional[PensionParameters]:
    if load_workbook is None:
        return None
    try:
        wb = load_workbook(filename=path, data_only=True)
    except Exception:
        return None

    indexation: Dict[int, float] = {}
    indexation_main: Dict[int, float] = {}
    indexation_sub: Dict[int, float] = {}
    salaries: Dict[int, float] = {}
    inflation: Dict[int, float] = {}
    minimum_pension: Optional[float] = None
    sick_male: Optional[float] = None
    sick_female: Optional[float] = None

    for ws in wb.worksheets:
        title = (ws.title or "").lower()
        parsed = _parse_year_value_sheet(ws)
        if not parsed:
            # try single cell parameters
            try:
                # search for keywords in the first column
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    key = str(row[0]).lower() if row[0] is not None else ""
                    if not key:
                        continue
                    if ("min" in key and "emeryt" in key) or ("minimal" in key and "emeryt" in key):
                        for cell in row[1:]:
                            if isinstance(cell, (int, float)):
                                minimum_pension = float(cell)
                                break
                    if "chorob" in key and "m" in key:  # male
                        for cell in row[1:]:
                            if isinstance(cell, (int, float)):
                                sick_male = float(cell)
                                break
                    if "chorob" in key and ("k" in key or "kobiet" in key or "female" in key):
                        for cell in row[1:]:
                            if isinstance(cell, (int, float)):
                                sick_female = float(cell)
                                break
            except Exception:
                pass
            continue

        if "walory" in title or "index" in title or "waloryz" in title:
            indexation.update(parsed)
        elif "wynagrod" in title or "sred" in title or "śred" in title:
            salaries.update(parsed)
        elif "inflac" in title:
            inflation.update(parsed)
        elif "emeryt" in title and ("min" in title or "minimal" in title):
            # might be a single-year table too
            # pick the last value as current minimum
            if parsed:
                minimum_pension = list(sorted(parsed.items()))[-1][1]

    params = PensionParameters(
        indexation_rates=indexation if indexation else {
            2020: 0.0356, 2021: 0.0447, 2022: 0.0724, 2023: 0.1412, 2024: 0.1288
        },
        average_salaries=salaries if salaries else {
            2023: 7155.48, 2024: 7850.00
        },
        inflation_rates=inflation if inflation else {
            2023: 0.112, 2024: 0.035
        },
        minimum_pension=minimum_pension if minimum_pension is not None else 1588.44,
        sick_leave_avg_days_male=sick_male if sick_male is not None else 15.2,
        sick_leave_avg_days_female=sick_female if sick_female is not None else 21.5,
    )
    return params


# Funkcja pomocnicza do inicjalizacji kalkulatora z domyślnymi parametrami
def create_default_calculator():
    """Tworzy kalkulator z parametrami z pliku 'Parametry-III 2025.xlsx' jeśli dostępny, w przeciwnym razie używa domyślnych."""

    # Spróbuj wczytać parametry z pliku Excel
    excel_path = os.path.join(os.path.dirname(__file__), "Parametry-III 2025.xlsx")
    params_from_excel = load_parameters_from_excel(excel_path)

    if params_from_excel is not None:
        return PensionCalculator(params_from_excel)

    # Fallback: Przykładowe dane makroekonomiczne (jak wcześniej)
    params = PensionParameters(
        indexation_rates={
            2020: 0.0356,
            2021: 0.0447,
            2022: 0.0724,
            2023: 0.1412,
            2024: 0.1288,
            2025: 0.0562  # prognoza
        },
        average_salaries={
            2023: 7155.48,
            2024: 7850.00,
            2025: 8400.00  # prognoza
        },
        inflation_rates={
            2023: 0.112,
            2024: 0.035,
            2025: 0.045  # prognoza
        }
    )

    return PensionCalculator(params)


# API dla frontendu (Flask endpoint example)
def calculate_pension_api(request_data: dict) -> dict:
    """
    Endpoint API dla kalkulacji emerytury

    Przykład request_data:
    {
        "age": 35,
        "gender": "M",
        "gross_salary": 8000,
        "work_start_year": 2012,
        "planned_retirement_year": 2055,
        "include_sick_leave": true,
        "ppk_contribution_rate": 0.02,
        "ppk_employer_rate": 0.015
    }
    """

    try:
        # Walidacja danych wejściowych
        if not all(k in request_data for k in ["age", "gender", "gross_salary",
                                               "work_start_year", "planned_retirement_year"]):
            return {"error": "Brakuje wymaganych pól"}

        # Przygotuj dane użytkownika
        user_input = UserInput(
            age=request_data["age"],
            gender=Gender(request_data["gender"]),
            gross_salary=request_data["gross_salary"],
            work_start_year=request_data["work_start_year"],
            planned_retirement_year=request_data["planned_retirement_year"],
            account_balance=request_data.get("account_balance"),
            subaccount_balance=request_data.get("subaccount_balance"),
            include_sick_leave=request_data.get("include_sick_leave", True),
            salary_history=request_data.get("salary_history"),
            postal_code=request_data.get("postal_code"),
            ppk_contribution_rate=request_data.get("ppk_contribution_rate", 0),
            ppk_employer_rate=request_data.get("ppk_employer_rate", 0),
            ikze_annual_amount=request_data.get("ikze_annual_amount", 0)
        )

        # Utwórz kalkulator i oblicz
        calculator = create_default_calculator()
        result = calculator.calculate(user_input)

        # Dodaj metadane
        result["calculation_date"] = datetime.now().isoformat()
        result["user_age"] = request_data["age"]
        result["user_gender"] = request_data["gender"]

        return result

    except Exception as e:
        return {"error": str(e)}


# Funkcja do generowania ciekawostek
def get_random_fact():
    """Zwraca losową ciekawostkę o emeryturach"""
    facts = [
        "Najwyższą emeryturę w Polsce (ponad 40 000 zł) otrzymuje były górnik z Katowic, który przepracował 48 lat.",
        "Średnia emerytura kobiet jest o 33% niższa niż mężczyzn ze względu na krótszy okres składkowy.",
        "Co 10 Polak odkłada dodatkowe środki na emeryturę w ramach III filaru.",
        "Najload_parameters_from_excelniższe emerytury występują w województwie podkarpackim, najwyższe w śląskim.",
        "Tylko 15% Polaków korzysta z PPK mimo automatycznego zapisu.",
        "Opóźnienie emerytury o 5 lat może zwiększyć świadczenie nawet o 40%.",
        "W 2060 roku na jednego emeryta przypadać będzie tylko 1.2 osoby pracującej."
    ]

    import random
    return random.choice(facts)


# Przykład użycia
if __name__ == "__main__":
    # Test kalkulatora
    test_input = UserInput(
        age=35,
        gender=Gender.MALE,
        gross_salary=8000,
        work_start_year=2012,
        planned_retirement_year=2055,
        include_sick_leave=True,
        ppk_contribution_rate=0.02,
        ppk_employer_rate=0.015
    )

    calculator = create_default_calculator()
    result = calculator.calculate(test_input)

    print("=== Wyniki kalkulacji emerytury ===")
    print(f"Emerytura nominalna: {result['pension_nominal']} zł")
    print(f"Emerytura realna (siła nabywcza): {result['pension_real']} zł")
    print(f"Stopa zastąpienia: {result['replacement_rate']}%")
    print(f"Porównanie do średniej: {result['comparison_to_average']}%")
    print(f"Całkowity kapitał: {result['total_capital']} zł")
